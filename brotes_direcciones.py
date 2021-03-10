#!/usr/bin/env python
# coding: utf-8

# In[217]:


import pandas as pd
from textpack import tp
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from xlsxwriter.utility import xl_range
from shutil import copyfile
# from styleframe import StyleFrame
import sys

def agrupar_similares(path, col='DIRECCIÓN', out='./planilla_madre_brotes_direcciones 27.xlsx'):
    madre = pd.read_excel(path,engine="openpyxl")
    print('working')

    # ## Grouping by Levenshtein distance and TF-IDF score

    # In[218]:


    def clean_direcciones(df):
        df['FIS'] = pd.to_datetime(df['FIS'], format='%Y-%m-%d', errors='coerce')
        remove_words = ['avenida', 'calle', 'pasaje','av','av\.','pasaje','psj', 'psj\.', "depto", "departamente", 'casa']
        pat = r'\b(?:{})\b'.format('|'.join(remove_words))
        df['direccion_limpia'] = df[col].str.replace(pat, '',case=False)
        df['direccion_limpia']  = df['direccion_limpia'].str.normalize('NFKD').str.encode('ascii', errors='ignore')        .str.decode('utf-8')        .str.rstrip().to_frame()
        return df

    def group_by_tf_id_score(df, col):
        groups = tp.TextPack(madre, [col])
        groups.run()
        groups.export_csv('./temp-grouped.csv')
        df = pd.read_csv('./temp-grouped.csv')
        os.remove('./temp-grouped.csv')
        return df
        
        
    madre_grouped = group_by_tf_id_score(clean_direcciones(madre), 'direccion_limpia')
    madre_grouped.insert(madre_grouped.columns.get_loc('DIRECCIÓN')+1,'BROTE','')
    madre_grouped.insert(madre_grouped.columns.get_loc('BROTE')+1,'OBSERVACION BROTE','')
    madre_grouped


    # df['FIS']## Filter by date

    # In[219]:


    def filter_by_date(df, start_date, end_date=datetime.today()):
        df['FIS'] = pd.to_datetime(df['FIS'], format='%Y-%m-%d', errors='coerce')
        in_date_range = (df['FIS'] >= start_date) & (df['FIS'] <= end_date)
        mask = in_date_range
        df = df[mask].copy()
        return df

    # madre_grouped_filteres_by_date = filter_by_date(madre_grouped, datetime(2021,1, 23))
    madre_grouped_filteres_by_date = madre_grouped
    print(madre_grouped_filteres_by_date)
    # madre_grouped_filteres_by_date = madre_grouped_filteres_by_date.sort_values(by=[col])
    # madre_grouped_filteres_by_date = madre_grouped_filteres_by_date.sort_values(by=['freq'] , ascending=False)
    # madre_grouped_filteres_by_date = madre_grouped_filteres_by_date.sort_values(by=['Group'])
    comunas_df = {
        'antofagasta': madre_grouped_filteres_by_date[madre_grouped_filteres_by_date['COMUNA'] == 'ANTOFAGASTA'],
        'calama': madre_grouped_filteres_by_date[madre_grouped_filteres_by_date['COMUNA'] == 'CALAMA'],
        'mejillones': madre_grouped_filteres_by_date[madre_grouped_filteres_by_date['COMUNA'] == 'MEJILLONES'],
        'tocopilla': madre_grouped_filteres_by_date[madre_grouped_filteres_by_date['COMUNA'] == 'TOCOPILLA'],
        'taltal': madre_grouped_filteres_by_date[madre_grouped_filteres_by_date['COMUNA'] == 'TALTAL'],
        'sanpedro': madre_grouped_filteres_by_date[madre_grouped_filteres_by_date['COMUNA'] == 'SAN PEDRO DE ATACAMA'],
        'sierragorda': madre_grouped_filteres_by_date[madre_grouped_filteres_by_date['COMUNA'] == 'SIERRA GORDA'],
        'ollague': madre_grouped_filteres_by_date[madre_grouped_filteres_by_date['COMUNA'] == 'OLLAGUE'],
        'mariaelena': madre_grouped_filteres_by_date[madre_grouped_filteres_by_date['COMUNA'] == 'MARIA ELENA'],
    }
    # ## Apply colors

    # In[220]:


    def apply_color_by_group(df):
        background_colors = [
            'background-color: #45BBCC',
            'background-color: #819699',
            'background-color: #70FFC2',
            'background-color: #FFB0B0',
            'background-color: #CC4589',
            'background-color: #854ECC',
            'background-color: #8F8799',
            'background-color: #7B89FF',
            'background-color: #FFE9BA',
            'background-color: #978799',
            'background-color: #FFA28E',
            'background-color: #7AFFF5',
            'background-color: #BDBB39',
            'background-color: #8A896E',
            'background-color: #F0D160',
            'background-color: #9DBBF2',
            'background-color: #398ABD',
            'background-color: #F29DC5',
            'background-color: #C29B58',
            'background-color: #8A806E',
            'background-color: #F09460',
            'background-color: #9DF2D9',
            'background-color: #8ABFA0',
            
        ]
        
        color_dict = {}
        for i,address_group in enumerate(list(df['Group'].unique())):
            color_dict[address_group] = background_colors[i% len(background_colors)]
            i+=1

        def apply_color(row):
            return [color_dict[row['Group']]]*len(row)

        return df.style.apply(apply_color,axis=1)
    writer = pd.ExcelWriter('temp.xlsx', datetime_format='dd/mm/yyyy')
    for key in comunas_df.keys():
        comunas_df[key]['freq'] = comunas_df[key].groupby('Group')['Group'].transform('count')
        comunas_df[key]['Group'] = comunas_df[key]['Group'] + comunas_df[key]['freq'].map(str)
        comunas_df[key] = comunas_df[key].sort_values(by=['Group', 'freq'], ascending=False)
        madre_styled = apply_color_by_group(comunas_df[key])
        madre_styled.to_excel(writer, sheet_name=key, index=False, engine='openpyxl'),
    writer.save()
    wb = load_workbook('temp.xlsx')

    wb.save('planila-brotes-domiciliarios_04-03.xlsx')




agrupar_similares("C:/seremi/brotes/reporte brotes domiciliarios/04-03/CONFIRMADOS + PROBABLES.xlsx")